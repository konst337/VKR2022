from flask import Flask, request, render_template, send_from_directory
import re
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter


app = Flask(__name__)

@app.route("/model", methods=['GET', 'POST'])
def model():
    mapp = readFile('map.xlsx')
    map1 = readMap(mapp)
    if request.method == 'GET':
        return render_template('model.html', map1=map1)
    else:
        data = request.form.get('techs')
        plats = request.form.get('platform')
        if data == "":
            return render_template('model.html', map1=map1)
        req, plats = parseReq(data, plats)     
        answ = readAttack(req,plats)
        return render_template('res.html',answ=answ)


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/load')
def download():
    document_path = os.getcwd()
    try:
        return send_from_directory(document_path, "mapped-techniques.xlsx")
    except:
        return render_template('res.html')


def readAttack(req,plats): # функция формирования итоговой таблицы
    technics = readFile('enterprise-attack-v11.2-techniques.xlsx') 
    mapp = readFile('map.xlsx')
    mittab = readFile('enterprise-attack-v11.2-mitigations.xlsx') 
    arr = []
    tactics = readMap(mapp)

    for i in req:
        tab1 = mapp.loc[mapp['Techniques'] == i]
        techs = tab1.iloc[0]['Map']
        techs = techs.split(" ")       
        for j in techs: # поиск соответствия техник
            tab2 = technics.loc[technics['ID MITRE'] == j]
            mitig = findMitigations(j,mittab)
            for keys,values in tactics.items():
                result = re.findall(r'([TТ]\d*\.\d*)', " ".join(values))
                if tab1.iloc[0]['Techniques'] in result:
                    tab1.iloc[0]['Tactics'] = keys
                    break
            if tab2.empty: # если нет соответствия
                if not plats:
                    arr.append(tab1)
            elif plats == "": # если нет фильтра
                result = tab1.join(tab2, how="cross")
                result.insert(13, "Mitigations", mitig)
                arr.append(result)
            else: 
                platforms = tab2.iloc[0]['platforms']
                platforms = platforms.split(",") 
                platforms = list(map(str.strip, platforms))
                if plats in platforms:
                    result = tab1.join(tab2, how="cross")
                    result.insert(13, "Mitigations", mitig)
                    arr.append(result)
    if arr:  
        new_df = pd.concat(arr)
        new_df.drop(columns = ['Map','created','last modified','version','is sub-technique','sub-technique of','defenses bypassed','contributors','impact type','relationship citations','system requirements', 'CAPEC ID', 'supports remote'],axis = 1, inplace=True)
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(new_df, index=False, header=True):
            ws.append(r)
        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        temp = ws['A'][1].value
        k = 1                   
        for i in range(2, len(ws['A'])-1): # поиск и объединение одинаковых ячеек тактик
            if temp == ws['A'][i].value and temp != ws['A'][i+1].value:
                ws.merge_cells(start_row=k+1, start_column=1, end_row=i+1, end_column=1)
                k = i
                temp = ws['A'][i].value
            elif temp != ws['A'][i].value:
                k = i
                temp = ws['A'][i].value 
            elif ws['A'][i].value == ws['A'][i+1].value and i == len(ws['A'])-2:
                ws.merge_cells(start_row=k+1, start_column=1, end_row=i+2, end_column=1)
                break                

        for row in ws.iter_rows():
            for cell in row:      
                cell.alignment =  cell.alignment.copy(wrapText=True,horizontal="center",vertical="top") 
                if isinstance(cell.value, str):
                    cell.value = cell.value.replace("<code>","")
                    cell.value =  cell.value.replace("</code>","")

        for i in range(1, ws.max_column+1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = 15
            if ws[letter][0].value in ['Description','description','detection',"Mitigations"]:
                ws.column_dimensions[letter].width = 25
            ws[letter][0].value = ws[letter][0].value.upper()
            ws[letter][0].alignment = ws[letter][0].alignment.copy(vertical="center") 

        wb.save("mapped-techniques.xlsx")
        
        return 1
    else:
        return 0

def readMap(data): # функция считывания таблицы соответствий
    map1 = {}
    techs = []
    current = ""
    for i in data.itertuples():
        if not pd.isnull(i[1]):
            current = i[1]
            techs = []
            techs.append(i[2] + "   " + i[3])          
        else:
            techs.append(i[2] + "   " + i[3])   
            map1[current] = techs

    return map1
    
def parseReq(data, plats): # функция обработки запроса на фильтрацию
    result = re.findall(r'([TТ]\d*\.\d*)', data)
    platforms = ['Azure AD', 'Containers', 'Google Workspace', 'IaaS', 'Linux', 'Network', 'Office 365', 'PRE', 'SaaS', 'Windows', 'macOS']
    for i in platforms:
        if plats.lower() in i.lower():
            plats = i
            break
    if plats == 'any':
        plats = ""

    return result,plats

def readFile(file): # функция считывания файла
    document_path = os.getcwd()+'/'+file 
    excel_data = pd.read_excel(document_path)
    data = pd.DataFrame(excel_data)
    return data

def findMitigations(tech, mitig): # поиск способов нейтрализации
    tab2 = mitig.loc[mitig['target ID'] == tech]
    res = ''
    if tab2.empty:
        return res

    for i, row in tab2.iterrows():
        sourc = row.loc['source ID']
        descr = row.loc['mapping description']
        res += "{}: {}\n".format(sourc, descr)

    return res

